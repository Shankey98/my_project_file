# dart_api_module.py

import os
import zipfile
import requests
import pandas as pd
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter


# ------------------------
# 1. 기업 고유코드 다운로드 및 검색
# ------------------------
def download_corp_code_file(api_key: str, output_path: str = "./CORPCODE_2024.csv") -> pd.DataFrame:
    """
    DART API로부터 CORPCODE.xml 파일을 받아서 DataFrame으로 반환
    """
    zip_url = f"https://opendart.fss.or.kr/api/corpCode.xml?crtfc_key={api_key}"
    zip_file = "CORPCODE.zip"
    xml_file = "CORPCODE.xml"

    response = requests.get(zip_url)
    if b"<result>" in response.content:
        raise Exception("API 호출 실패: 인증키 오류 또는 기타 문제")

    with open(zip_file, 'wb') as f:
        f.write(response.content)

    with zipfile.ZipFile(zip_file, 'r') as zip_ref:
        zip_ref.extractall(".")

    tree = ET.parse(xml_file)
    root = tree.getroot()
    company_data = []
    for item in root.findall('list'):
        company = {child.tag: child.text for child in item}
        company_data.append(company)

    df = pd.DataFrame(company_data)
    df = df[df['stock_code'] != ' '].copy()
    df['corp_code'] = df['corp_code'].astype(str).str.zfill(8)
    df['stock_code'] = df['stock_code'].astype(str).str.zfill(6)
    df['corp_name'] = df['corp_name'].str.replace(' ', '')
    df.to_csv(output_path, index=False)
    return df

def get_corp_code(company_name: str, df: pd.DataFrame):
    matched = df[df['corp_name'].str.contains(company_name)]
    if matched.empty:
        raise ValueError(f"기업명을 찾을 수 없습니다: {company_name}")

    print("\n[검색 결과: 유사 기업명 리스트]")
    for idx, row in matched.reset_index().iterrows():
        print(f"{idx}: {row['corp_name']} (종목코드: {row['stock_code']})")

    selected = int(input("\n위 목록 중 원하는 기업의 번호를 입력하세요: "))
    selected_row = matched.reset_index().iloc[selected]
    return selected_row['corp_code'], selected_row['corp_name']



# ------------------------
# 2. 공시 재무정보 요청
# ------------------------
def fetch_financial_data_by_years(api_key: str, corp_code: str, corp_name: str, bs_years: list, reprt_code: str = '11011', fs_div: str = 'C') -> pd.DataFrame:
    data_list = []
    for year in bs_years:
        url = f"https://opendart.fss.or.kr/api/fnlttSinglAcntAll.json?crtfc_key={api_key}&corp_code={corp_code}&bsns_year={year}&reprt_code={reprt_code}&fs_div={fs_div}"
        response = requests.get(url)
        data = response.json()
        if 'list' in data:
            for item in data['list']:
                data_list.append({
                    '기업코드': corp_code,
                    '기업명': corp_name,
                    '사업년도': item['bsns_year'],
                    '재무제표종류': item['sj_div'],
                    '계정명': item['account_nm'],
                    '당기계정과목': item['thstrm_nm'],
                    '당기계정금액': item['thstrm_amount'],
                    '전기계정과목': item['frmtrm_nm'],
                    '전기계정금액': item['frmtrm_amount']
                })
    return pd.DataFrame(data_list)


# ------------------------
# 3. 예시 실행 코드 (테스트용) 및 엑셀로 추출.
# ------------------------
if __name__ == "__main__":
    API_KEY = "4f9a5b61c0d015b864f8143ef2d8c233fb93e263"
    COMPANY_NAME = input("기업명을 입력하세요: ").strip()
    BS_YEARS = list(range(2023, 2025))
    fs_div_options = {
    "CFS": "연결재무제표",
    "OFS": "별도재무제표"
    }
    print("\n[재무제표 구분 선택]")
    for code, name in fs_div_options.items():
        print(f"{code}: {name}")
    FS_DIV = input("위 코드 중 하나를 입력하세요 (예: CFS): ").strip().upper()

    report_options = {
    "11011": "사업보고서",
    "11012": "반기보고서",
    "11013": "1분기보고서",
    "11014": "3분기보고서"
    }
    print("\n[보고서 유형 선택]")
    for code, name in report_options.items():
        print(f"{code}: {name}")
    REPRT_CODE = input("위 코드 중 하나를 입력하세요 (예: 11011): ").strip()

    corp_df = download_corp_code_file(API_KEY)
    corp_code, selected_name = get_corp_code(COMPANY_NAME, corp_df)
    FS_DIV_TYPE = "연결" if FS_DIV == "CFS" else "별도" #OFS는 별도

    df_rslt = fetch_financial_data_by_years(API_KEY, corp_code, selected_name, BS_YEARS, reprt_code=REPRT_CODE, fs_div=FS_DIV)

    if not df_rslt.empty:
        for col in ['당기계정금액', '전기계정금액']:
            if col in df_rslt.columns:
                df_rslt[col] = df_rslt[col].str.replace(',', '').replace('', '0').astype(float)


    output_filename = f"{selected_name}_{FS_DIV_TYPE}재무제표_모음.xlsx"
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        meta_df = pd.DataFrame({
            "항목": ["기업명", "사업연도 (범위)", "재무제표 구분", "보고서 코드"],
            "값": [selected_name, f"{min(BS_YEARS)}~{max(BS_YEARS)}", FS_DIV_TYPE, REPRT_CODE]
        })
        meta_df.to_excel(writer, sheet_name="정보", index=False)

        for sj, group_df in df_rslt.groupby("재무제표종류"):
            sheet_name = sj[:31]
            group_df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(output_filename)
    for sheet in wb.sheetnames:
        if sheet == "정보":
            continue
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = max_length + 5

        header = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell, col_name in zip(row, header):
                if col_name in ['당기계정금액', '전기계정금액'] and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'

    wb.save(output_filename)
    print(df_rslt.head())
    print(f"\n 엑셀 파일로 저장 완료: {output_filename}")
